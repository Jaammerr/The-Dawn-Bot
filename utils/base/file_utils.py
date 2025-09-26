import asyncio
import time
import aiofiles

from pathlib import Path
from loguru import logger

from openpyxl.reader.excel import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from models import ModuleType, OperationResult



class FileOperations:
    def __init__(self, base_path: str = "./results"):
        self.base_path = Path(base_path)
        self.lock = asyncio.Lock()
        self.module_paths: dict[ModuleType, dict[str, Path]] = {
            "tasks": {
                "success": self.base_path / "tasks" / "tasks_success.txt",
                "failed": self.base_path / "tasks" / "tasks_failed.txt",
            },
            "stats": {
                "base": self.base_path / "stats" / "accounts_stats.xlsx",
            },
            "accounts": {
                "unlogged": self.base_path / "accounts" / "unlogged_accounts.txt",
                "invalid_proxy": self.base_path / "accounts" / "invalid_proxy_accounts.txt",
            },
            "login": {
                "success": self.base_path / "login" / "login_success.txt",
                "failed": self.base_path / "login" / "login_failed.txt",
            },
        }

    async def setup_files(self):
        self.base_path.mkdir(exist_ok=True)
        for module_name, module_paths in self.module_paths.items():
            for path_key, path in module_paths.items():
                path.parent.mkdir(parents=True, exist_ok=True)

                if module_name == "stats":
                    continue
                else:
                    path.touch(exist_ok=True)

    async def setup_stats(self):
        self.base_path.mkdir(exist_ok=True)

        for module_name, module_paths in self.module_paths.items():
            if module_name == "stats":
                timestamp = int(time.time())
                for path_key, path in module_paths.items():
                    path.parent.mkdir(parents=True, exist_ok=True)
                    if path_key == "base":
                        new_path = path.parent / f"accounts_stats_{timestamp}.xlsx"
                        self.module_paths[module_name][path_key] = new_path
                        await asyncio.to_thread(self._create_excel_with_header, new_path)

    @staticmethod
    def _create_excel_with_header(xlsx_path: Path, sheet_name: str = "Stats"):
        wb = Workbook()
        ws: Worksheet = wb.active
        ws.title = sheet_name
        ws.append([
            "Email",
            "Email Password",
            "Points",
            "Referral Code",
            "Total Referrals",
            "Total Referral Points",
            "Completed Tasks",
        ])

        widths = [46, 66, 14, 14, 14, 20, 24]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        wb.save(xlsx_path)

    async def _append_excel_row(self, xlsx_path: Path, row: list, sheet_name: str = "Stats"):
        await asyncio.to_thread(self._append_excel_row_sync, xlsx_path, row, sheet_name)

    @staticmethod
    def _append_excel_row_sync(xlsx_path: Path, row: list, sheet_name: str = "Stats"):
        wb = load_workbook(xlsx_path)
        ws: Worksheet = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
        ws.append(row)
        wb.save(xlsx_path)
        wb.close()

    async def export_result(self, result: OperationResult, module: ModuleType):
        if module not in self.module_paths:
            raise ValueError(f"Unknown module: {module}")

        file_path = self.module_paths[module][
            "success" if result["status"] else "failed"
        ]
        async with self.lock:
            try:
                async with aiofiles.open(file_path, "a") as file:
                    await file.write(f"{result['email']}:{result['email_password']}\n")
            except IOError as e:
                logger.error(f"Account: {result['email']} | Error writing to file (IOError): {e}")
            except Exception as e:
                logger.error(f"Account: {result['email']} | Error writing to file: {e}")


    async def export_invalid_account(self, email: str, password: str | None, reason: str):
        if reason not in self.module_paths["accounts"]:
            raise ValueError(f"Unknown reason: {reason}")

        file_path = self.module_paths["accounts"][reason]
        async with self.lock:
            try:
                async with aiofiles.open(file_path, "a") as file:
                    if password:
                        await file.write(f"{email}:{password}\n")
                    else:
                        await file.write(f"{email}\n")
            except IOError as e:
                logger.error(f"Account: {email} | Error writing to file (IOError): {e}")
            except Exception as e:
                logger.error(f"Account: {email} | Error writing to file: {e}")


    async def export_invalid_proxy_account(self, email: str, password: str | None, proxy: str):
        if "invalid_proxy" not in self.module_paths["accounts"]:
            raise ValueError("Invalid proxy path not found in module paths")

        file_path = self.module_paths["accounts"]["invalid_proxy"]
        async with self.lock:
            try:
                async with aiofiles.open(file_path, "a") as file:
                    if password:
                        await file.write(f"{email}:{password}:{proxy}\n")
                    else:
                        await file.write(f"{email}:{proxy}\n")
            except IOError as e:
                logger.error(f"Account: {email} | Error writing to file (IOError): {e}")
            except Exception as e:
                logger.error(f"Account: {email} | Error writing to file: {e}")

    async def export_stats(self, result: OperationResult):
        file_path = self.module_paths["stats"]["base"]
        async with self.lock:
            try:
                if result["status"] is True:
                    row = [
                        result["email"],
                        result["email_password"],
                        result["data"]["user_info"]["points"],
                        result["data"]["referral_stats"]["referralCode"],
                        result["data"]["referral_stats"]["totalReferrals"],
                        result["data"]["referral_stats"]["totalPointsEarned"],
                        False,
                    ]
                else:
                    row = [
                        result["email"],
                        result["email_password"],
                        "N/A",
                        "N/A",
                        "N/A",
                        "N/A",
                        "N/A",
                    ]

                await self._append_excel_row(file_path, row)

            except IOError as e:
                logger.error(f"Account: {result['email']} | Error writing to Excel (IOError): {e}")
            except Exception as e:
                logger.error(f"Account: {result['email']} | Error writing to Excel: {e}")
